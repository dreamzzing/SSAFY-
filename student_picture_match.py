import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

# 현재 작업 경로
current_dir = os.getcwd()

# 엑셀파일 찾기
excel_file = None
for file in os.listdir(current_dir):
    if file.endswith(".xlsx") and not file.startswith("~$"):  # 임시 파일 제외
        excel_file = file
        break
if not excel_file:
    print("엑셀 파일을 찾을 수 없습니다. 현재 폴더에 .xlsx 파일을 추가하세요.")
    input("엔터를 입력하여 종료하세요")
    exit()

# 사진폴더 찾기
pictures_dir = None
for folder in os.listdir(current_dir):
    if os.path.isdir(os.path.join(current_dir, folder)):
        pictures_dir = os.path.join(current_dir, folder)
        break
if not pictures_dir:
    print("사진 폴더를 찾을 수 없습니다. 현재 폴더에 사진 폴더를 추가하세요.")
    input("엔터를 입력하여 종료하세요")
    exit()

# 파일명을 위한 기수 설정
coach_th = input("기수를 입력하세요(숫자만 입력하시오) : ")

# 지역명 설정
local_name = input("지역 이름을 입력하세요 : ")

# 엑셀 파일 읽기
df = pd.read_excel(excel_file)

# 숫자와 . 제거
df['공통트랙'] = df['공통트랙'].str.replace(r'^\d\.\s*', '', regex=True)

# 지역 필터링 (입력받은 지역만 추출)
df = df[df['지역'] == local_name].copy()

# 동명이인 체크
df['비고'] = None  # 비고란 추가
name_counts = {}
for name in df['이름']:
    if name in name_counts:
        name_counts[name] += 1
    else:
        name_counts[name] = 1

# 동명이인 비고 작성
for index, row in df.iterrows():
    if name_counts[row['이름']] > 1:
        df.at[index, '비고'] = '동명이인 있음'

# 이름과 사진 경로를 매칭하기 위한 딕셔너리 생성
picture_dict = {}
for picture in os.listdir(pictures_dir):
    if picture.endswith(".jpg"):
        # 파일명에서 이름 추출 (형식: 대전_강건호5742.jpg)
        name = picture.split("_")[1].rsplit(".")[0].rstrip("0123456789")
        picture_dict[name] = os.path.join(pictures_dir, picture)

# 데이터프레임 컬럼 순서 변경
df_new = pd.DataFrame()
df_new['이름'] = df['이름']
df_new["사진"] = None
df_new['학번'] = df['학번']
df_new['학과'] = df['전공']  # 전공 컬럼을 학과로 사용
df_new['트랙(공통)'] = df['공통트랙']  # 공통트랙 컬럼을 트랙(공통)으로 사용
df_new['1학기 전공'] = df['1학기 트랙']  # 1학기트랙을 1학기 전공으로 사용
df_new['1학기 반'] = df['분반\n(1학기)']  # 분반(1학기) 컬럼을 1학기 반으로 사용
df_new['비고'] = df['비고']  # 비고란 추가
df_new['공통 분반'] = df['공통분반']


def insert_images_to_excel(df, output_file):
    # 기존 파일 삭제
    if os.path.exists(output_file):
        os.remove(output_file)

    # 데이터프레임을 엑셀로 저장
    df.to_excel(output_file, index=False, engine='openpyxl')

    # 워크북 로드
    wb = load_workbook(output_file)
    ws = wb.active
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:  # 비어 있지 않은 셀만 처리
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 각 행에 맞는 이미지를 첨부
    for index, row in df.iterrows():
        name = row["이름"]
        if name in picture_dict:
            image_path = picture_dict[name]
            img = Image(image_path)

            # 이미지 크기 조정 (가로 70, 세로는 비율 유지)
            original_width, original_height = img.width, img.height
            img.width = 100
            img.height = original_height * (img.width / original_width)

            # 이미지 삽입 (엑셀 행의 2번째 열에 해당)
            cell_address = f"B{index + 2}"  # +2는 헤더와 0-based index를 고려한 값
            ws.add_image(img, cell_address)

            # 행 높이 조정 (이미지 높이에 맞춤)
            ws.row_dimensions[index + 2].height = img.height * 0.75  # 약간의 여유 추가

    # 열 너비 조정
    ws.column_dimensions['B'].width = 15

    # 엑셀 저장
    wb.save(output_file)
    
    
# 새로운 엑셀 파일에 데이터 저장
output_file = f"{coach_th}기_{local_name}_교육생_명단.xlsx"

try:
    # 결과 저장 및 이미지 첨부
    insert_images_to_excel(df_new, output_file)
    print(f"작업 완료. 결과가 {output_file}에 저장되었습니다.")
    input("엔터를 눌러 종료합니다.")
except Exception as e:
    print(f"오류가 발생했습니다: {e}")
    input("오류 발생! 엔터를 눌러 종료합니다.")  # 오류 메시지 확인 후 종료 대기
